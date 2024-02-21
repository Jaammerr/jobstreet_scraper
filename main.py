import asyncio
import uuid
import pyuseragents
import noble_tls

from openpyxl import Workbook
from typing import NamedTuple, List
from loguru import logger


class JobData(NamedTuple):
    title: str
    classification: str
    company_name: str
    location: str
    salary: str
    work_type: str


class TestScraper(noble_tls.Session):
    def __init__(self, search_term: str):
        super().__init__()

        self.headers = {
            'authority': 'www.jobstreet.com.my',
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'en-US,en;q=0.9,ru;q=0.8',
            'referer': 'https://www.jobstreet.com.my/microsoft-jobs',
            'user-agent': pyuseragents.random(),
        }
        self.client_identifier = "chrome_111"
        self.random_tls_extension_order = True
        self.timeout_seconds = 10

        self.search_params = {
            'siteKey': 'MY-Main',
            'sourcesystem': 'houston',
            'userqueryid': '93f6fca1cd53a0bf7ee91020256bbed2-2962548',
            'page': '1',
            'seekSelectAllPages': 'true',
            'keywords': search_term,
            'pageSize': '30',
            'include': 'seodata',
            'locale': 'en-MY',
            'solId': uuid.uuid4().hex,
        }

    @staticmethod
    def raise_for_status(response):
        if response.status_code != 200:
            raise Exception(f"Request failed with status code {response.status_code}")



    async def setup_session(self):
        response = await self.get("https://www.jobstreet.com.my/")
        self.raise_for_status(response)

        session_id = response.cookies.get("JobseekerSessionId")
        if not session_id:
            logger.error("No session id found")
            exit(1)

        self.search_params["usersessionid"] = session_id
        self.search_params["userid"] = session_id
        self.search_params["eventCaptureSessionId"] = session_id

    @staticmethod
    def extract_jobs_data(jobs: List[dict]) -> List[JobData]:
        jobs_data = []

        for job in jobs:
            title = job.get("title")
            classification = job.get("classification").get("description")
            company_name = job.get("companyName")
            location = job.get("jobLocation")
            full_location = f"{location.get('label')} ({location.get('countryCode')})"
            salary = job.get("salary")
            work_type = job.get("workType")

            jobs_data.append(JobData(title, classification, company_name, full_location, salary, work_type))

        return jobs_data

    async def search_jobs(self) -> List[JobData]:
        url: str = "https://www.jobstreet.com.my/api/chalice-search/v4/search"

        response = await self.get(url, params=self.search_params)
        self.raise_for_status(response)

        jobs = response.json().get('data')
        if not jobs:
            logger.warning(f"No available jobs found for {self.search_params['keywords']}")

        else:
            logger.info(f"Found {len(jobs)} jobs for {self.search_params['keywords']}")
            return self.extract_jobs_data(jobs)

    @staticmethod
    def export_to_excel(jobs: List[JobData]) -> None:
        try:
            workbook = Workbook()
            sheet = workbook.active

            headers = list("Title, Classification, Company Name, Location, Salary, Work Type".split(", "))
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col, value=header)

            for row, job in enumerate(jobs, start=2):
                for col, field in enumerate(job, start=1):
                    sheet.cell(row=row, column=col, value=field)

            workbook.save("jobs.xlsx")
            logger.success("Jobs data exported to jobs.xlsx")

        except Exception as error:
            logger.error(f"Error exporting to excel: {error}")
            exit(1)

    async def start(self) -> None:
        try:
            await self.setup_session()
            jobs = await self.search_jobs()
            self.export_to_excel(jobs)

        except Exception as error:
            logger.error(f"Failed to process scraper: {error}")
            exit(1)


if __name__ == "__main__":
    word = input("Enter a search term: ")
    scraper = TestScraper(word)
    asyncio.run(scraper.start())
